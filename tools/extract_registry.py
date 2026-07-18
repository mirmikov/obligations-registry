#!/usr/bin/env python3
"""Extract the source workbook into deterministic JSON used by the Go seed loader."""

from __future__ import annotations

import json
import sys
import re
import zipfile
import xml.etree.ElementTree as ET
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


FIELDS = [
    "account_type",
    "entry_date",
    "counterparty",
    "legal_entity",
    "cost_category",
    "priority",
    "responsible",
    "document_number",
    "deferment_days",
    "document_date",
    "amount",
    "planned_payment_date",
    "approval_date",
    "actual_payment_date",
    "status",
    "urgency",
    "comment",
]

REFERENCE_COLUMNS = OrderedDict(
    [
        ("statuses", 1),
        ("cost_categories", 2),
        ("priorities", 3),
        ("urgencies", 4),
        ("legal_entities", 5),
        ("responsibles", 8),
        ("account_types", 9),
    ]
)


def clean(value):
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def unique(values):
    result = []
    seen = set()
    for value in values:
        value = clean(value)
        if value in (None, "") or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: extract_registry.py <source.xlsx> <output.json>")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])

    values_book = load_workbook(source, data_only=True, read_only=False)
    sheet = values_book["Реестр"]
    dictionaries = values_book["Справочники"]

    notes_by_row = {}
    with zipfile.ZipFile(source) as archive:
        threaded = ET.fromstring(archive.read("xl/threadedComments/threadedComment1.xml"))
        for comment in threaded:
            match = re.fullmatch(r"H(\d+)", comment.attrib.get("ref", ""))
            texts = ["".join(node.itertext()).strip() for node in comment if node.tag.endswith("text")]
            if match and texts:
                notes_by_row[int(match.group(1))] = "\n".join(text for text in texts if text)

    records = []
    source_row = 1
    for row_number in range(2, sheet.max_row + 1):
        values = [clean(sheet.cell(row_number, col).value) for col in range(1, 18)]
        # A filled-down formula in column L must not turn an otherwise blank row
        # into a business record. At least one source field must contain data.
        source_values = values[:11] + values[12:]
        if not any(value not in (None, "") for value in source_values):
            continue
        source_row += 1
        record = dict(zip(FIELDS, values))
        record["source_note"] = notes_by_row.get(row_number)
        record["source_row"] = row_number
        records.append(record)

    references = {}
    for key, column in REFERENCE_COLUMNS.items():
        references[key] = unique(
            dictionaries.cell(row, column).value for row in range(2, dictionaries.max_row + 1)
        )

    references["account_types"] = [value for value in references["account_types"] if value != "Все"]

    references["counterparties"] = unique(record["counterparty"] for record in records)

    payload = {
        "source": source.name,
        "sheet": "Реестр",
        "headers": [sheet.cell(1, col).value for col in range(1, 18)],
        "records": records,
        "references": references,
        "source_rules": {
            "planned_payment_date": "Дата документа + Отсрочка дней",
            "overdue": "Плановая дата оплаты раньше даты анализа, статус не Оплачено и не Отменено",
            "due_soon_days": 3,
        },
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Extracted {len(records)} records to {output}")


if __name__ == "__main__":
    main()
